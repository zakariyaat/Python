'''																								
print("Zakariya Ahmed.")																									
																									
print("Hi \nThis is Tamim.")																									
																									
print('*' * 10)																									
																									
price = 25																									
rating = 4.9																									
name = 'ABCD'																									
is_published = True																									
print(rating)																									
print(f"{name}  ${price} {rating}")																									
																									
m = 2.5																									
j = 5																									
print(m+j)																									
																									
full_name = 'John Smith'																									
age = 20																									
status = 'new'																									
is_new = True																									
print(full_name, age, status)																									
																									
name = input('What is your name?')																									
colour = input('What is your favourite colour?')																									
print('Hi ' + name)																									
print(name + ' likes ' + colour)																									
																									
name = input("Enter your name: ")																									
age = input("Enter your age: ")																									
print("Hello " + name + "! you are " + age)																									
																									
first_name = 'Zakariya'																									
middel_name = 'Ahmed'																									
last_name = 'Tamim'																									
print("This is", first_name, middel_name, last_name, ".")																									
print(first_name, "Ahmed", last_name)																									
																									
birth_year = input('Birth year: ')																									
age = 2024 - int(birth_year)																									
print(type(age))																									
print(type(birth_year))																									
print(age)																									
																									
weight_lbs = input('Weight (lbs): ')																									
weight_kg = .4535 * int(weight_lbs)																									
print(str(weight_kg) + ' kilograms')																									
print(f"{weight_kg} kilograms")																									
																									
courses = "Python's for beginners"																									
course = 'Python course for "Beginners"'																									
print(courses, course)																									
																									
																									
mail = ''																									
Hi Tamim,																									
																									
Here is our first email to you.																									
																									
Thank you,																									
The support team																									
''         # Triple quotes for multiline string input																									
print(mail)																									
																									
#STRINGS																									
																									
course = 'Python for Beginners'																									
print(course[0])																									
print(course[1:5])																									
print(course[-4])																									
print(course[:6])																									
print(course[6:])																									
print(course[:])																									
print(course[1:-1])																									

#freeCodeCamp.org (Intermediate Python Programming Course)
#Strings: ordered, immutable, text representarion
my_string = "I'm a programmer"   # my_string = 'I\'m a programmer'
print(my_string)
multiLineString = """Hi!
This is Tamim.
"""
print(multiLineString)

s = "Hello World"
s1 = s[::2]
print(s1)
s2 = s[::-1]
print(s2)

s = "     Hello"
print(s)
gretting = s.strip()
print(gretting)
name = "Tom"
sentence = gretting + " " + name + "!"
print(sentence)

for x in gretting:
    print(x)
if 'ello' in gretting:
    print('Yes')
else:
    print('No')

print(sentence.upper())
print(sentence.lower())
print(sentence.startswith('Hello'))
print(sentence.endswith('World'))
print(sentence.find('o'))
print(sentence.find('ello'))
print(sentence.find('p'))
print(sentence.count('o'))
print(sentence.count('p'))
print(sentence.replace('Tom', 'World'))

string = 'How are you doing ?'
list = string.split()
print(list)
s = 'How,are,you,doing'
l =s.split(",")
print(l)
new_string = ''.join(string)
print(new_string)
new_list = ' '.join(list)
print(new_list)
n_s = ''.join(s)
print(n_s)

my_list = ['a']*6
print(my_list)
my_string = ''.join(my_list)
print(my_string)

from timeit import default_timer as timer
my_list = ['a'] *1000000
#bad
start = timer()
my_string = ''
for i in my_list:
    my_string += i
stop = timer()
print(stop - start)
#good
start = timer()
my_string = ''.join(my_list)
stop = timer()
print(stop - start)


#%, .foma(), f-Strings  # %s for integer, %d for integer,%f for floting point
var = "Tom"
x = 5
pi =3.14159
my_string = "the variable is %s" %var 
print(my_string)
my_integer = "the variable is %d" %x
print(my_integer)
my_fp = "the variable is %.2f" %pi  #my_fp = "the variable is {}".format(pi)
print(my_fp)
w = 5.62
my_fp_nums = "the variable is {:.2f} and {}".format(pi,w)
print(my_fp_nums)
my_fp_numbs = f"the variable is {pi} and {w*2.5}"
print(my_fp_numbs)


#FORMATTED STRINGS																									
																									
first = 'John'																									
last = 'Smith'																									
message = first + ' [' + last + '] is a coder'																									
msg = f'{first} [{last}] is a coder'																									
print(first, '[', last, '] is a coder')																									
print(message)																									
print(msg)																									
																									
#STRING METHODS																									
																									
course = 'Python for Beginners.'																									
print(len(course))																									
print(course.upper())																									
print(course)																									
print(course.lower())																									
print(course.find('f'))																									
print(course.find('s'))																									
print(course.find('w'))																									
print(course.find('x'))																									
print(course.replace('Beginners', 'Absolute Beginners'))																									
print(course.replace('P', 'J'))																									
print('Python' in course)																									
print('python' in course)																									
print('For' in course)																									
print('for' in course)																									
																									
#ARITHMERIC OPERATION																									
																									
print(10 + 3)																									
print(10 - 3)																									
print(10 * 3)																									
print(10 / 3)																									
print(10 // 3)																									
print(10 % 3)																									
print(10 ** 3)																									
print(2 ** 5)																									
																									
x = 10																									
x = x + 3																									
print(x)																									
x += 2																									
print(x)																									
y = 10																									
y -= 3																									
print(y)																									
																									
num1 = input("Enter a number: ")																									
num2 = input("Enter another number: ")																									
summ = num1 + num2																									
print(summ)																									
result = float(num1) + float(num2)																									
print(result)																									
																									
#OPERATION PRECEDENCE																									
																									
x = 10 + 3 * 2																									
print(x)  #Operation Order 0.Bracket 1.exponentiation(**) 2.multiplication/division 3.addition/subtraction																									
y = 10 + 3 * 2 ** 3																									
print(y)																									
y = (10 + 3) * 2 ** 3																									
print(y)																									
																									
x = 2.9																									
print(round(x))																									
print(abs(-x))																									
																									
#MATH FUNCTION																									
																									
import math																									
print(math.ceil(2.9))																									
print(math.floor(2.9))																									
																									
import math																									
x = 5																									
y = 2																									
z = 6																									
i = 'j'																									
print(math.comb(x, y))    #comb( n, k)= n! /(k! * (n-k)!) {Return the number of ways to choose k items from n items without repetition and without order.}																									
print(math.factorial(z))																									
print(math.isfinite(x))																									
print(math.isinf(y))    #Return True if x is a positive or negative infinity, and False otherwise.																									
print(math.isnan(z))    #Return True if x is a NaN (not a number), and False otherwise.																									
																									
import math																									
print(math.isqrt(49))																									
print(math.lcm(8, 12, 16))																									
print(math.modf(9.5))																									
																									
#Power and logarithmic functions																									
																									
print(math.exp(1))																									
print(math.cbrt(27))																									
print(math.exp2(1))																									
print(math.log(1000, 10))																									
print(math.log1p(10)) #math.log1p(x) =ln(x) =Return the natural logarithm of 1+x (base e)																									
print(math.log10(1000)) #more accurate than log(x, 10)																									
print(math.pow(4, 2))																									
																									
#Constants																									
import math																									
print(math.pi)																									
print(math.e)																									
print(math.tau)																									
print(-math.inf)																									
																									
#Trigonometric functions																									
import math																									
print("math.sin(30) =", math.sin(.523599))    #python takes input as radian & gives output as radian ****																									
print("math.sin(60) =", math.sin(1.0472))																									
print("math.tan(45) =", math.tan(.7853987))																									
print("math.cos(90) =", math.cos(1.5708))																									
print("math.sin(90) =", math.sin(1.5708))																									
																									
x = math.radians(60)                     #math.radians(x) Convert angle x from degrees to radians																									
y = math.degrees(1.0472)                 #Convert angle x from radians to degrees																									
print("math.cos(x) =", math.cos(x))     #python takes input as radian ****																									
print("math.sin(x) =", math.sin(x))																									
print("math.tan(x) = ", math.tan(x))																									
print("math.cos(y) =", math.cos(y))      #python takes input as radian & gives output as radian ****																									
print("math.cos(60) =", math.cos(60))    #python takes input & output as radian ****																									
print("math.acos(.5) =", math.acos(.5))    #acos = Arc Cos/ Inverse cos																									
																									
																									
#CONDITIONS																									
																									
is_hot = False																									
is_cold = False																									
if is_hot:																									
    print("It is a hot day!!")																									
    print("Drink plenty water.")																									
elif is_cold:																									
    print("It's a cold day.")																									
    print("Wear warm clothes.")																									
else:																									
    print("It's a lovely day.")																									
print("Enjoy your day.")																									
																									
																									
price = 1000000																									
has_good_credit = False																									
if has_good_credit:																									
    down_payment = 0.1 * price																									
else:																									
    down_payment = 0.2 * price																									
print(f"Need to put down: ${down_payment} ")																									
																									
																									
has_high_income = True																									
has_good_credit = True																									
if has_high_income and has_good_credit:																									
    print("Eligible for loan.")																									
else:																									
    print("Not eligible for loan.")																									
																									
has_high_income = True																									
has_good_credit = False																									
if has_high_income or has_good_credit:																									
    print("Eligible for loan.")																									
else:																									
    print("Not eligible for loan.")																									
																									
																									
has_good_credit = True																									
has_criminal_record = True																									
if has_good_credit and not has_criminal_record:																									
    print("Eligible for loan.")																									
else:																									
    print("Not eligible for loan.")																									
																									
temperature = int(input("temperature: "))																									
if temperature >= 30:																									
    print("It's a hot day.")																									
elif temperature <= 15:																									
    print("It's a cold day.")																									
else:																									
    print("It's neither hot nor cold.")																									
print("Have a good day.Enjoy your day.")																									
																									
name = "Zakariya Ahmed"																									
size = len(name)																									
if size < 4:																									
    print("Name must be at least 4 characters.")																									
elif size > 20:																									
    print("Name must be a maximum of 20 characters.")																									
else:																									
    print("Name looks good.")																									
																									
																									
weight = int(input("Weight: "))																									
unit = input("(L)bs or (K)g: ")																									
if unit.upper() == "K":																									
    converted = 2.2 * weight																									
    print(f"You are {converted} pounds.")																									
else:																									
    converted = .45 * weight																									
    print(f"You are {converted} kilos.")																									
																									
																									
num1 = float(input("Enter first number: "))																									
op = input("Enter operator: ")																									
num2 = float(input("Enter second number: "))																									
																									
if op == "+":																									
    print(num1 + num2)																									
elif op == "-":																									
    print(num1 - num2)																									
elif op == "/":																									
    print(num1 / num2)																									
elif op == "*":																									
    print(num1 * num2)																									
else:																									
    print("Invalid operator!!")																									
																									
#While Loop																									
																									
i = 1																									
while i <= 5:																									
    print(i)																									
    i = i+1																									
print("Done")																									
																									
i = 1																									
while i <= 5:																									
    print("*" * i)																									
    i = i+1																									
print("Done")																									
																									
																									
secret_word = "giraffe"																									
gues = ""																									
while gues != secret_word:																									
    gues = input("Enter guess: ")																									
																									
print("You win!")																									
																									
																									
secret_number = 9																									
guess_count = 0																									
guess_limit = 3																									
while guess_count < guess_limit:																									
    guess = int(input("Guess: "))																									
    guess_count += 1																									
    if guess == secret_number:																									
        print("You won!")																									
    break																									
else:																									
    print("Sorry, you failed!!")																									
																									
																									
secret_word = "Tamim"																									
guess = ""																									
guess_count = 0																									
guess_limit = 5																									
out_of_guesses = False																									
while guess != secret_word and not (out_of_guesses):																									
    if guess_count < guess_limit:																									
        guess = input("Enter guess: ")																									
        guess_count += 1																									
else:																									
    out_of_guesses = True																									
																									
if out_of_guesses:																									
    print("Out of guesses, YOU LOSE")																									
else:																									
    print("You win!")																									
																									
																									
# Problem ***																									
																									
#command = ""																									
started = False																									
while True:																									
    command = input(">").lower()																									
    if command == "start":																									
        if started:																									
            print("Car is already started!")																									
        else:																									
            started = True																									
            print("Car started...")																									
    elif command == "stop":																									
        if not started:																									
            print("Car is already stopped!")																									
        else:																									
            started = False																									
            print("Car stopped.")																									
    elif command == "help":																									
        print("""																									
        start - to start the car																									
        stop  - to stop the car																									
        quit  - to quit																									
        """)																									
    elif command == "quit":																									
        break																									
    else:																									
        print("Sorry, I don't understand that!")																									


#For Loop																									


for item in 'Python':																									
    print(item)																									
for item in ['Python', 'Javascript', 'c++']:																									
    print(item)																									
for item in [1, 2, 3, 4, 5, 6]:																									
    print(item)																									
for item in range(6):																									
    print(item)																									
for item in range(4, 10):																									
    print(item)																									
for item in range(1, 10, 2):																									
    print(item)																									
																									
																									
prices = [10, 20, 30]																									
total = 0																									
for price in prices:																									
    total += price																									
print(f"Total: {total}")																									
																									
																									
def rais_to_power(base_num, pow_num):																									
    result = 1																									
    for index in range(pow_num):																									
        result = result * base_num																									
        return result																									
																									
base = int(input("Enter Base: "))																									
power = int(input("Enter Power: "))																									
print(rais_to_power(base, power))																									
																									
																									
#Nested For Loop																									
																									
for x in range(4):																									
    for y in range(3):																									
        print(f'({x}, {y})')																									
																									
#Problem																									
																									
n = [5, 2, 5, 2, 2]																									
for y in n:																									
    print('X' * y)																									
																									
numbers = [5, 2, 5, 2, 2]																									
for x_count in numbers:																									
    output = ""																									
    for count in range(x_count):																									
        output += 'x'																									
print(output)																									
																									
x = ''																									
for y in range(5):																									
    x += 'X'																									
print(x)																									
																									
numbers = [12, 12, 3, 5, 5, 3, 12, 12]																									
for x_count in numbers:																									
    output = ""																									
for count in range(x_count):																									
    output += 'x'																									
print(output)																									
																									
#LISTS																									
																									
names = ['John', 'Bob', 'Mosh', 'Sarah', 'Mary']																									
print(names)																									
print(names[2])																									
print(names[-2])																									
print(names[2:-1])																									
print(names[3:])																									
names[0] = 'Jon'																									
print(names)																									
																									
numbers = [3, 6, 2, 4, 10, 8, 5]																									
max = numbers[0]																									
for number in numbers:																									
    if number > max:																									
        max = number																									
print(f"Maximum number is {max}")																									
																									
#2D LISTS																									
																									
matrix = [																									
[1, 2, 3],																									
[4, 5, 6],																									
[7, 8, 9]																									
]																									
print(matrix[0][1])																									
print(matrix[2])																									
print(matrix[1][1])																									
matrix[1][0] = 40																									
print(matrix[1])																									
																									
matrix = [																									
[1, 2, 3],																									
[4, 5, 6],																									
[7, 8, 9]																									
]																									
for row in matrix:																									
    for item in row:																									
        print(item)																									
																									
#LIST METHODS																									
																									
numbers = [5, 2, 1, 7, 4]																									
print(numbers)																									
numbers.append(20)																									
print(f"numbers after 20 is added at last  {numbers}")																									
numbers.insert(2, 10)																									
print(f"numbers after 10 is added at 3rd position or index 2 {numbers}")																									
numbers.insert(0, 10)																									
print(f"numbers after 10 is added at 1st position or index 0 {numbers}")																									
numbers.remove(5)																									
print(f"numbers after removed 5 {numbers}")																									
print(numbers.pop())      #Last number																									
numbers.pop()																									
print(f"numbers after removing last number {numbers}")																									
print(numbers.count(10))  #counting same number in list																									
print(numbers.index(7))   #Index/Existance of a number																									
print(3 in numbers)       #Existance of a number																									
numbers.sort()																									
print(numbers)																									
numbers.clear()																									
print(numbers)																									
																									
numbers = [5, 2, 7, 4, 6, 7, 8, 1, 2, 9, 3, 2, 5, 2, 6, 5, 7, 9, 4]																									
print(numbers.count(2))																									
print(numbers.count(7))																									
print(5 in numbers)																									
numbers.reverse()																									
print(numbers)																									
numbers = [5, 2, 7, 4, 6, 7, 8, 1, 2, 9, 3, 2, 5, 2, 6, 5, 7, 9, 4]																									
numbers.sort()																									
print(numbers)																									
numbers = [5, 2, 7, 4, 6, 7, 8, 1, 2, 9, 3, 2, 5, 2, 6, 5, 7, 9, 4]																									
numbers.sort()																									
numbers.reverse()																									
print(numbers)																									
																						
numbers = [5, 2, 7, 4, 6, 7, 8, 1]																								
numbers2 = numbers.copy()																									
numbers.reverse()																									
numbers.append(3)																									
print(numbers2)																									
print(numbers)																									
																									
lucky_numbers = [4, 8, 15, 16, 23, 42]																									
friends = ["Kevin", "Karen", "Jim", "Oscar", "Tby"]																									
friends.extend(lucky_numbers)																									
print(friends)																									
friends.insert(1, "Kally")																									
print(friends)


#freeCodeCamp.org (Intermediate Python Programming Course)
mylist = [4, 3, 1, -1, -5, 10]
new_list = sorted(mylist)
print(new_list)
print(mylist)
total_list = mylist + new_list
print(total_list)

list = [1,2,3,4,5,6,7,8,9]
print(list)
a = list[:4]
print(a)
b = list[5:]
print(b)
c = list[0:6:2]
print(c)
d = list[::-1]
print(d)

list1 = [1] * 5
print(list1)

list_org = ["banana", "cherry","apple"]
list_copy = list_org.copy()     #or list_copy = list_org[:]
list_copy.append("lemon")
print(list_copy)
print(list_org)

x = [1,2,3,4,5,6]
y = [i*i for i in x]
print(x)
print(y)
for i in x:

mytuple = tuple(["Max", 28, "Boston"])
print(mytuple)
for i in mytuple:
    print(i)

#Unpack Tuple
my_tuple = "Max", 28, "Boston"
name, age, city = my_tuple
print(name)
print(age)
print(city)

my_tuple1 = (0,1,2,3,4)
i1,*i2,i3 = my_tuple1
print(i1)
print(i3)
print(i2)

import sys
my_list = [0, 1, 2,"hello", True]
my_tuple = (0, 1, 2,"hello", True)
print(sys.getsizeof(my_list), "bytes")
print(sys.getsizeof(my_tuple), "bytes")

import timeit 
print(timeit.timeit(stmt="[0,1,2,3,4,5,]", number= 1000000))
print(timeit.timeit(stmt="(0,1,2,3,4,5,)", number= 1000000))

#end freeCodeCamp.org


#Removing duplicates																									
numbers = [2, 2, 4, 6, 3, 4, 6, 1]																									
uniques = []																									
for number in numbers:																									
    if number not in uniques:																									
        uniques.append(number)																									
print(uniques)																									
																									
#TUPLES																									
																									
numbers = (1, 2, 3)   #using () instade of [] cann't modify the list even cann't change the value of index																									
print(numbers.count(3))			 # numbers = 1,2,3 (also a tuples)																		
print(numbers.index(3))																									
																									
#Unpacking																									
																									
coordinates = (1, 2, 3)																									
x = coordinates[0]																									
y = coordinates[1]																									
z = coordinates[2]																									
print(x, y, z)																									
																									
coordinates = (1, 2, 3, 4)																									
w, x, y, z = coordinates																									
print(w, x, y, z)																									
																									
coordinates = [1, 2, 3, 4, 5]																									
v, w, x, y, z = coordinates																									
print(v, w, x, y, z)																									
																									
#DICTIONARIES																									
																									
customer = {																									
"name": "John Smith",																									
"age": 30,																									
"is_verified": True																									
}																									
print(customer["name"])																									
print(customer.get("birthdate"))																									
print(customer.get("age"))																									
print(customer.get("birthdate","Jun 5 1980"))																									
																									
customer["name"] = "Jack Smith"																									
print(customer["name"])																									
customer["birthdate"] = "Jan 1 1990"																									
print(customer["birthdate"])																									
print(customer["name"])																									
																									
																									
phone = input("phone Number: ")																									
digits_mapping = {																									
"1": "One",																									
"2": "Two",																									
"3": "Three",																									
"4": "Four",																									
"5": "Five",																									
"6": "Six",																									
"7": "Senem",																									
"8": "Eight",																									
"9": "Nine",																									
"0": "Zero"																									
}																									
output = ""																									
for ch in phone:																									
    output += digits_mapping.get(ch, "!") + " "																									
print(output)       #phone: 12a34   >>  One Two ! Three Four																									
																									
#Emoji Converter																									
																									
message = input(">")																									
words = message.split(' ')																									
print(words)																									
emojis = {																									
":)": "😇",																									
":(": "😢"																									
}																									
output = ""																									
for word in words:																									
    output += emojis.get(word, word) + " "																									
print(output)																									


#freeCodeCamp.org (Intermediate Python Programming Course)
mydict = {"name":"Max", "age":28, "city": "New York"}
print(mydict)
mydict["email"] = "mac@xyz.com"
print(mydict)
for key in mydict:
    print(key)
for value in mydict:
    print(value)
for key, value in mydict.items():
    print(key, value)

mydict_copy = mydict.copy()      # mydict_copy = dict(mydict)
mydict_copy["phone"] = "+088 966 56 6435 9785 "
print(mydict_copy)
print(mydict)

mydict = {"name":"Max", "age":28, "email":"max@xyz.com"}
mydict_2 = dict(name="Mary", age=27, city="Boston")
mydict.update(mydict_2)
print(mydict)


#freeCodeCamp.org (Intermediate Python Programming Course)
#Sets (unorderd, mutable, no duplicates)
myset = {1,2,4,3,2,1}
print(myset)   #Does not alow duplicates to print & prints in orbitary order 
for i in myset:
    print(i)    
if 4 in myset:
    print("yes")
   
set1 = set([1,2,3])
print(set1)
set2 = set("Hello!")
print(set2)

set0 = set()
print(type(set0))
set0.add(4)
set0.add(5)
set0.add(9)
set0.add(6)
set0.add(7)
print(set0)
set0.remove(9)
print(set0)
set0.discard(7)
print(set0)
print(set0.pop())
print(set0)

odds ={1,3,5,7,9}
evens = {0,2,4,6,8}
primes = {2,3,5,7}
u = odds.union(evens)
print(u)
i = odds.intersection(evens)
print(i)
pio = odds.intersection(primes)
print(pio)
pie = evens.intersection(primes)
print(pie)

setA = {1,2,3,4,5,6,7,8,9}
setB = {1,2,3,10,11,12}
diff = setA.symmetric_difference(setB)
print(diff)
diff1 = setA.difference(setB)
print(diff1)
diff2 = setB.difference(setA)
print(diff2)
#setA.intersection_update(setB)
#print(setA)
setA.difference_update(setB)
print(setA)
setA.update(setB)
print(setA)

setA = {1,2,3,4,5,6}
setB = {1,2,3}
setC = {7,8}
print(setA.issubset(setB))
print(setB.issubset(setA))

print(setA.issuperset(setB))
print(setB.issuperset(setA))

print(setA.isdisjoint(setB))
print(setB.isdisjoint(setC))

setD = setA.copy()   #setD = set(setA)
setD.add(7)
print(setD)
print(setA)

a = frozenset([1,2,3,4])
print(a)     #a.add(2) or a.remove(4)   will be error!!


#freeCodeCamp.org (Intermediate Python Programming Course)
#Collections: Counter, namedtuple, OrderedDict, deque
from collections import Counter
a = "aaaaabbbbccc"
my_counter = Counter(a)
print(my_counter)
print(my_counter.items())
print(my_counter.keys())
print(my_counter.values())
print(my_counter.most_common(1))
print(my_counter.most_common(2))
print(my_counter.most_common(1)[0][0])
print(list(my_counter.elements()))

from collections import namedtuple
point = namedtuple('point', 'x,y')
pt = point(1,-4)
print(pt.x, pt.y)

from collections import OrderedDict
OrderedDict = OrderedDict()
OrderedDict['a'] = 1
OrderedDict['b'] = 2
OrderedDict['c'] = 3
OrderedDict['d'] = 4
print(OrderedDict)

from collections import defaultdict
d = defaultdict(int)
d['a'] = 1
d['b'] = 2
print(d['a'])
print(d['c'])
e = defaultdict(float)   #int,float,list...
e['a'] = 1
e['b'] = 2
print(e['b'])
print(e['c'])

from collections import deque
d = deque()

d.append(1)
d.append(2)
d.append(3)

d.appendleft(3)
print(d)

d.pop()
print(d)

d.rotate(-1)
print(d)

d.extend([4,5,6])
print(d)

d.extendleft([9,8,7])
print(d)

d.rotate(2)
print(d)

#end freeCodeCamp go to >>intermediate.py


#FUNCTIONS																									
																									
																									
def greet_user():																									
    print("Hi there!")																									
    print("Welcome aboard.")																									
																									
																									
print("Start")																									
greet_user()																									
print("Finish.")																									
																									
																									
def say_hi(name, age):																									
    print("Hello " + name + ", you are " + str(age) + ".")																									
																									
say_hi("Mike", 35)																									
say_hi("Steve", 70)																									
																									
#PARAMETERS																									
																									
																									
def greet_user(name):																									
    print(f"Hi {name}!")																									
    print("Welcome aboard.")																									
																									
																									
print("Start")																									
greet_user('John')																									
greet_user('Mary')																									
print("Finish.")																									
																									
																									
def greet_user(first_name, last_name):																									
    print(f"Hi {first_name} {last_name}!")																									
    print("Welcome aboard.")																									
																									
																									
print("Start")																									
greet_user('John', 'Smith')																									
print("Finish.")																									
																									
#Keyword Arguments																									
																									
																									
def greet_user(first_name,last_name):																									
    print(f"Hi {first_name} {last_name}!")																									
    print("Welcome aboard.")																									
																									
																									
print("Start")																									
greet_user(last_name='John', first_name='Smith')																									
print("Finish.")																									
																									
#Return Statement																									
																									
																									
def square(number):																									
    return number * number																									
																									
																									
result = square(3)																									
print(result)																									
print(square(5))																									
																									
																									
def square(number):																									
    print(number * number)																									
																									
																									
result = square(3)																									
print(result)																									
print(square(5))																									
																									
																									
def square(number):																									
    print(number * number)																									
    return number * number																									
																									
																									
result = square(3)																									
print(result)																									
print(square(5))																									
																									
#Creating a Reusable Function																									
																									
																									
def emoji_converter(message):																									
    words = message.split(' ')																									
    emojis = {																									
        ":)": "😊",																									
        ":(": "😒"																									
    }																									
    output = ""																									
    for word in words:																									
        output += emojis.get(word, word) + " "																									
        return output																									
																									
																									
message = input(">")																									
result = emoji_converter(message)																									
print(result)																									
																									
#EXCEPTION																									
																									
try:																									
    age = int(input('Age:'))																									
    print(age)																									
except ValueError:																									
    print('Invalid value')																									
																									
																									
try:																									
    age = int(input('Age:'))																									
    income = 200000																									
    risk = income / age																									
    print(age)																									
except ZeroDivisionError:																									
    print('Age cannot be 0.')																									
except ValueError:																									
    print('Invalid value')																									
																									
#CLASSES																									
																									
class Point:																									
    def move(self):																									
        print("move")																									
																									
    def draw(self):																									
        print("draw")																									
																									
																									
point1 = Point()																									
point1.x = 10																									
point1.y = 20																									
print(point1.x)																									
point1.draw()																									
																									
print2 = Point()																									
print2.y = 2																									
print(print2.y)																									
																									
#CONSTRUCTORS																									
																									
class Point:																									
    def __init__(self,x,y):																									
        self.x = x																									
        self.y = y																									
																									
def move(self):																									
    print("move")																									
																									
def draw(self):																									
    print("draw")																									
																									
																									
point = Point(10, 20)																									
print(point.x)																									
print(point.y)																									
point.x = 11																									
print(point.x)																									
																									
																									
class Person:																									
    def __init__(self, name):																									
        self.name = name																									
																									
def talk(self):																									
    print(f"Hi, I am {self.name}")																									
																									
																									
john = Person("John Smith")																									
print(john.name)																									
john.talk()																									
																									
bob = Person("Bob Smith")																									
bob.talk()																									
																									
#INHERITANCE																									
																									
																									
class Mammal:																									
    def walk(self):																									
        print("Walk")																									
																									
																									
class Dog(Mammal):																									
    pass																									
class Cat(Mammal):																									
    pass																									
																									
																									
dog1 = Dog()																									
dog1.walk()																									
																									
																									
																									
class Mammal:																									
    def walk(self):																									
        print("Walk")																									
																									
																									
class Dog(Mammal):																									
    def bark(self):																									
        print("bark")																									
class Cat(Mammal):																									
    def be_annoying(self):																									
        print("annoying")																									
																									
																									
dog1 = Dog()																									
dog1.walk()																									
dog1.bark()																									
cat1 = Cat()																									
cat1.walk()																									
cat1.be_annoying()																									
																									
#MODULES																									
																									
import converters																									
																									
print(converters.kg_to_lbs(70))					    																				
print(converters.lbs_to_kg(140))					    																				
																									
																									
from converters import kg_to_lbs																									
print(kg_to_lbs(100))					                																				
from converters import lbs_to_kg																									
print(lbs_to_kg(200))																									
																									
																									
																									
import utils																									
																									
numbers =[10, 3, 6, 2]																									
max = utils.find_max(numbers)					    																				
print(max)					                            																			
from utils import find_max					            																				
numbers = [8, 3, 6, 2, 9]					            																				
print(f"Max = {find_max(numbers)}")																									
																									
numbers =[10, 6, 3, 2]																									
print(max(numbers))																									
																									
																									
#PACKAGES																									
																									
import ecommerce.shipping																									
ecommerce.shipping.calc_shipping()																									
ecommerce.shipping.calc_tex()																									
																									
from ecommerce.shipping import calc_shipping,calc_tex																									
calc_shipping()																									
calc_tex()																									
																									
																									
from ecommerce import shipping																									
shipping.calc_shipping()																									
shipping.calc_tex()																									
																									
#Generating Random Values																									
																									
import random																									
																									
for i in range(3):																									
    print(random.random())																									
																									
for i in range(3):																									
    print(random.randint(10, 20))																									
																									
																									
import random																									
																									
members = ['John', 'Mary', 'Bob', 'Mosh']																									
print(random.choice(members))																									
																									
																									
import random																									
																									
																									
class Dice:																									
    def roll(self):																									
        first = random.randint(1, 6)																									
        second = random.randint(1, 6)																									
        return (first, second)																									
																									
																									
dice = Dice()																									
print(dice.roll())																									
																									
#Files and Directories																									
																									
from pathlib import Path																									
#Absulate path (stored in C drive),  Relative path(stored in current directory/programm)																									
																									
path = Path("ecommerce")																									
print(path.exists())																									
																									
																									
																									
from pathlib import Path																									
																									
path = Path("ecommerce1")																									
print(path.exists())																									
																									
																									
																									
from pathlib import Path																									
																									
path = Path("emails")																									
print(path.exists())																									
																									
																									
from pathlib import Path																									
																									
path = Path("emails")																									
print(path.mkdir())																									
																									
																									
from pathlib import Path																									
																									
path = Path("emails")																									
print(path.rmdir())																									
																									
																									
																									
from pathlib import Path																									
																									
path = Path()																									
print(path.glob('*.*'))																									
																									
																									
from pathlib import Path																									
																									
path = Path()																									
print(path.glob('*.py'))																									
																									
																									
from pathlib import Path																									
																									
path = Path()																									
for file in path.glob('*.py'):																									
    print(file)																									
																									
																									
from pathlib import Path																									
																									
path = Path()																									
for file in path.glob('*'):																									
    print(file)																									
																									
																								
#Pypi & Pip																									
																									
																									
#Excel Spreadsheets																									
																									
import openpyxl as xl																									
from openpyxl.chart import BarChart, Reference																									
																									
# def process_workbook(filename):																									
wb = xl.load_workbook('transactions.xlsx')						import openpyxl as xl																			
sheet = wb['Sheet1']						from openpyxl.chart import BarChart, Reference																			
# cell = sheet['a1']																									
# cell = sheet.cell(1, 1)						def process_workbook(filename):																			
# print(cell.value)						wb = xl.load_workbook(filename)																			
						sheet = wb['Sheet1']																			
for row in range(1, sheet.max_row + 1):																									
print(row)																									
for row in range(2, sheet.max_row + 1):						for row in range(1, sheet.max_row + 1):																			
cell = sheet.cell(row, 3)						print(row)																			
# print(cell.value)						for row in range(2, sheet.max_row + 1):																			
corrected_price = cell.value * 0.9						cell = sheet.cell(row, 3)																			
corrected_price_cell = sheet.cell(row, 4)						# print(cell.value)																			
corrected_price_cell.value = corrected_price						corrected_price = cell.value * 0.9																			
						corrected_price_cell = sheet.cell(row, 4)																			
values = Reference(sheet,						corrected_price_cell.value = corrected_price																			
min_row=2,																									
max_row=sheet.max_row,						values = Reference(sheet,																			
min_col=4,						min_row=2,																			
max_col=4)						max_row=sheet.max_row,																			
						min_col=4,																			
chart = BarChart()						max_col=4)																			
chart.add_data(values)																									
sheet.add_chart(chart, 'e2')						chart = BarChart()																			
						chart.add_data(values)																			
wb.save('transactions.xlsx')						sheet.add_chart(chart, 'e2')																			
																									
						wb.save(filename)																			
'''
















